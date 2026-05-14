import json
import requests
import re
import base64
import os
import argparse
import smtplib
import time
from datetime import datetime
from typing import Dict, Tuple
from requests.auth import HTTPDigestAuth
from openai import OpenAI
import openpyxl
from openpyxl.drawing.image import Image
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import cv2
import numpy as np


# Thresholds (tune as needed)
DARK_THRESHOLD = 35              # pixel value below this is considered dark
DARK_PIXEL_RATIO = 0.98         # fraction of dark pixels to consider frame black
WHITE_THRESHOLD = 245           # mean above this -> white frame
STRIPE_VARIATION_RATIO = 0.995   # ratio of stddev of row/col sums to mean -> stripe
NOISE_LAPLACIAN_THRESH = 200     # minimal laplacian magnitude to count as high-frequency noise
NOISE_PIXEL_RATIO = 0.05           # fraction of pixels above NOISE_LAPLACIAN_THRESH to consider noisy

NG_CAMERAS = []
SYSTEM_ALERTS = []
ALL_HDD_STATUS = []

GMAIL_USER = "noreply.nivs@gmail.com"
GMAIL_APP_PASSWORD = "jjxpgndrikbftwed" 
GMAIL_TO = [
    "hai.tranviet@nidec.com",
    "phu.truongvan@nidec.com",
    "binh.lethai@nidec.com",
]

def is_black(img):
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    dark_ratio = np.count_nonzero(gray < DARK_THRESHOLD) / gray.size
    return dark_ratio >= DARK_PIXEL_RATIO


def is_white(img):
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    return np.mean(gray) > WHITE_THRESHOLD


def is_horizontal_stripe(img):
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    row_sums = gray.sum(axis=1).astype(np.float32)
    return (np.std(row_sums) / (np.mean(row_sums) + 1e-6)) > STRIPE_VARIATION_RATIO


def is_vertical_stripe(img):
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    col_sums = gray.sum(axis=0).astype(np.float32)
    return (np.std(col_sums) / (np.mean(col_sums) + 1e-6)) > STRIPE_VARIATION_RATIO

def is_noisy(img):
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    lap = cv2.Laplacian(gray, cv2.CV_64F)
    high_freq = np.abs(lap) > NOISE_LAPLACIAN_THRESH
    ratio = np.count_nonzero(high_freq) / high_freq.size
    return ratio > NOISE_PIXEL_RATIO

# ======================
# OVERLAY TEXT DETECTION (timestamp, camera label)
# ======================
def has_overlay_text(gray):
    h, w = gray.shape

    # vùng nghi ngờ có timestamp / camera name
    bottom = gray[int(h*0.80):, :]
    top = gray[:int(h*0.15), :]

    # detect edges
    edges_bottom = cv2.Canny(bottom, 30, 90)
    edges_top = cv2.Canny(top, 30, 90)

    # nếu vùng này có edge ratio cao => có text overlay
    if np.count_nonzero(edges_bottom) / edges_bottom.size > 0.002:
        return True
    if np.count_nonzero(edges_top) / edges_top.size > 0.002:
        return True

    return False


# DARK FRAME DETECTOR (has overlay)
# ======================
def is_black_or_no_detail(img):
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    mean_val = np.mean(gray)

    # Khung hình đen thật
    if mean_val < 10:
        return True

    lap = cv2.Laplacian(gray, cv2.CV_64F)
    sharpness = np.mean(np.abs(lap))

    # camera thật chỉ coi là 'no detail' khi <1.0
    return sharpness < 1.0
def has_structure(gray):
    h, w = gray.shape
    
    # tránh vùng chứa overlay text
    crop = gray[int(h*0.20):int(h*0.80), :]

    mean_val = np.mean(crop)
    if mean_val < 50:
        return False

    lap = cv2.Laplacian(crop, cv2.CV_64F)
    sharpness = np.mean(np.abs(lap))
    if sharpness < 2.0:
        return False

    edges = cv2.Canny(crop, 40, 120)
    ratio = np.count_nonzero(edges) / edges.size

    return ratio > 0.02

def is_too_dark(gray):
    return np.mean(gray) < 20      # NG ngay nếu ảnh tối toàn cục


def classify_image(path):
    img = cv2.imread(path)
    if img is None:
        return {"error": True, "description": "Cannot read image"}

    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

    # 1. Nếu ảnh quá tối → NG NGAY
    if is_too_dark(gray):
        return {"error": True, "description": "Ảnh đen sì"}

    # 3. Các lỗi còn lại → NG
    if is_black(img) or is_white(img) or is_horizontal_stripe(img) or \
       is_vertical_stripe(img) or is_noisy(img) or is_black_or_no_detail(img):
        return {"error": True, "description": ""}

    return {"error": False, "description": ""}


class CameraUtility:
    """Xử lý giao tiếp với camera hãng DAHUA"""
    @staticmethod
    def get_snapshot(base_url, username, password, channel_id, output_path, max_retries=2): # Chỉ cần thử 2 lần là đủ
        url = f"{base_url}/cgi-bin/snapshot.cgi?channel={channel_id+1}"
        
        for attempt in range(1, max_retries + 1):
            try:
                # Timeout giảm xuống một chút để phản ứng nhanh hơn
                response = requests.get(url, auth=HTTPDigestAuth(username, password), timeout=10)
                if response.status_code == 200:
                    with open(output_path, "wb") as f:
                        f.write(response.content)
                    return True
            except Exception as e:
                pass # Bỏ qua lỗi mạng cục bộ
                
            # Nếu chưa thành công, chỉ nghỉ đúng 2 giây (Không dùng lũy tiến nữa)
            if attempt < max_retries:
                # print(f"[*] Chụp thử lần {attempt} thất bại. Thử lại sau 2s...") # Tắt print cho đỡ rác màn hình
                time.sleep(2)
                
        # Thất bại chớp nhoáng -> Trả về False để Vòng 1 đẩy nó vào danh sách chờ 5 phút
        return False

    @staticmethod
    def get_channel_title_config(base_url, username, password):
        url = f"{base_url}/cgi-bin/configManager.cgi?action=getConfig&name=ChannelTitle"
        try:
            response = requests.get(url, auth=HTTPDigestAuth(username, password), timeout=10)
            if response.status_code == 200:
                data = response.text
                pattern = r"table\.ChannelTitle\[(\d+)\]\.Name=([^\n]+)"
                matches = re.findall(pattern, data)
                return {int(channel_id): name.strip() for channel_id, name in matches}
        except: pass
        return {}

    @staticmethod
    def get_hdd_status(base_url, username, password):
        url = f"{base_url}/cgi-bin/storageDevice.cgi?action=getDeviceAllInfo"
        try:
            response = requests.get(url, auth=HTTPDigestAuth(username, password), timeout=10)
            if response.status_code != 200 or "ErrorID" in response.text:
                return "Không hỗ trợ đọc HDD", False
                
            data = response.text
            # KIỂU 1
            indices_v1 = re.findall(r"table\.StorageDevice\[(\d+)\]", data)
            if indices_v1:
                max_idx = max([int(i) for i in indices_v1])
                results, has_error = [], False
                for i in range(max_idx + 1):
                    status_match = re.search(fr"StorageDevice\[{i}\]\.Status=(\w+)", data)
                    capacity_match = re.search(fr"StorageDevice\[{i}\]\.TotalBytes=(\d+)", data)
                    status = status_match.group(1) if status_match else "Unknown"
                    total_bytes = int(capacity_match.group(1)) if capacity_match else 0
                    total_tb = round(total_bytes / (1024**4), 2)
                    results.append(f"SATA-{i+1}: {status} ({total_tb}TB)")
                    if status not in ["Normal", "Active", "B.thường"]:
                        has_error = True
                return " | ".join(results), has_error

            # KIỂU 2
            indices_v2 = set(re.findall(r"list\.info\[(\d+)\]\.State", data))
            if indices_v2:
                results, has_error = [], False
                for idx in sorted(list(indices_v2)):
                    state_match = re.search(fr"list\.info\[{idx}\]\.State=(\w+)", data)
                    state = state_match.group(1) if state_match else "Unknown"
                    bytes_matches = re.findall(fr"list\.info\[{idx}\]\.Detail\[\d+\]\.TotalBytes=([\d\.]+)", data)
                    total_bytes = sum(float(b) for b in bytes_matches)
                    total_tb = round(total_bytes / (1024**4), 2) 
                    error_matches = re.findall(fr"list\.info\[{idx}\]\.Detail\[\d+\]\.IsError=(true)", data, re.IGNORECASE)
                    
                    if state != "Success" or error_matches:
                        has_error, status_text = True, "LỖI"
                    else:
                        status_text = "Bình thường"
                    results.append(f"Disk-{int(idx)+1}: {status_text} ({total_tb}TB)")
                return " | ".join(results), has_error

            return "Không có dữ liệu ổ cứng (No Disk)", True
        except Exception as e:
            return f"Lỗi lấy thông tin: {str(e)}", True

    @staticmethod
    def get_system_info(base_url, username, password):
        url = f"{base_url}/cgi-bin/magicBox.cgi?action=getSystemInfo"
        try:
            response = requests.get(url, auth=HTTPDigestAuth(username, password), timeout=10)
            if response.status_code == 200:
                ver = re.search(r"Version=([^\n]+)", response.text)
                return f"Online (Firmware: {ver.group(1).strip() if ver else 'N/A'})"
        except: pass
        return "Offline/Lỗi kết nối"
    

class HikvisionUtility:
    """Xử lý giao tiếp với camera hãng HIKVISION"""
    
    @staticmethod
    def get_snapshot(base_url, username, password, channel_id, output_path):
        ip_chan = channel_id + 32
        
        # --- CÁCH 1: DÙNG API ---
        urls_to_try = [
            f"{base_url}/ISAPI/Streaming/channels/{channel_id}02/picture",
            f"{base_url}/ISAPI/Streaming/channels/{channel_id}01/picture",
            f"{base_url}/ISAPI/Streaming/channels/{ip_chan}02/picture",
            f"{base_url}/ISAPI/Streaming/channels/{ip_chan}01/picture"
        ]
        for url in urls_to_try:
            try:
                response = requests.get(url, auth=HTTPDigestAuth(username, password), timeout=5)
                if response.status_code == 200 and len(response.content) > 1000:
                    with open(output_path, "wb") as f:
                        f.write(response.content)
                    return True
            except: 
                continue

        # --- CÁCH 2: DÙNG RTSP (Đã thêm cơ chế chờ Keyframe) ---
        ip_address = base_url.replace("http://", "").replace("https://", "").split(":")[0]
        rtsp_channels = [f"{channel_id}02", f"{channel_id}01", f"{ip_chan}02", f"{ip_chan}01"]
        
        for rtsp_chan in rtsp_channels:
            rtsp_url = f"rtsp://{username}:{password}@{ip_address}:554/Streaming/Channels/{rtsp_chan}"
            try:
                # Thêm cv2.CAP_FFMPEG để ép dùng bộ giải mã chuẩn
                cap = cv2.VideoCapture(rtsp_url, cv2.CAP_FFMPEG)
                if cap.isOpened():
                    # ĐỌC BỎ QUA 20 KHUNG HÌNH ĐẦU TIÊN ĐỂ CHỜ ẢNH NÉT (I-FRAME)
                    for _ in range(20):
                        ret, frame = cap.read()
                        if not ret:
                            break
                    
                    # Lưu lại khung hình đã ổn định
                    if ret and frame is not None:
                        cv2.imwrite(output_path, frame)
                        cap.release()
                        return True
                cap.release()
            except Exception:
                continue
                
        return False
    @staticmethod
    def get_channel_title_config(base_url, username, password):
        """Lấy danh sách các kênh, hỗ trợ nhiều đời NVR Hikvision và có cơ chế Failsafe"""
        channels = {}
        # Thử 2 đường dẫn API phổ biến nhất của Hikvision
        urls_to_try = [
            f"{base_url}/ISAPI/ContentMgmt/InputProxy/channels",
            f"{base_url}/ISAPI/System/Video/inputs/channels"
        ]
        
        for url in urls_to_try:
            try:
                response = requests.get(url, auth=HTTPDigestAuth(username, password), timeout=10)
                if response.status_code == 200:
                    # Đã sửa lỗi Regex ở đây
                    channel_blocks = re.findall(r'<(?:VideoInputChannel|InputProxyChannel)[^>]*>(.*?)</(?:VideoInputChannel|InputProxyChannel)>', response.text, re.DOTALL)
                    
                    for block in channel_blocks:
                        c_id = re.search(r'<id>(\d+)</id>', block)
                        c_name = re.search(r'<name>(.*?)</name>', block)
                        if c_id and c_name:
                            channels[int(c_id.group(1))] = c_name.group(1).strip()
                    
                    if channels:
                        return channels # Nếu lấy được thì trả về luôn
            except: 
                pass
                
        # --- CƠ CHẾ FAILSAFE ---
        print(f"[*] Không lấy được danh sách tên kênh API. Đang kích hoạt ép chụp ảnh thủ công...")
        for i in range(1, 17): 
            channels[i] = f"Camera {i:02d}"
            
        return channels

    @staticmethod
    def get_hdd_status(base_url, username, password):
        url = f"{base_url}/ISAPI/ContentMgmt/Storage"
        try:
            response = requests.get(url, auth=HTTPDigestAuth(username, password), timeout=20)
            if response.status_code == 200:
                hdd_blocks = re.findall(r'<hdd[^>]*>(.*?)</hdd>', response.text, re.DOTALL)
                if not hdd_blocks:
                    return "Không tìm thấy ổ cứng (No Disk)", True

                results, has_error = [], False
                for i, block in enumerate(hdd_blocks):
                    status_match = re.search(r'<status>(.*?)</status>', block, re.IGNORECASE)
                    capacity_match = re.search(r'<capacity>(.*?)</capacity>', block, re.IGNORECASE)
                    
                    status = status_match.group(1) if status_match else "Unknown"
                    total_mb = int(capacity_match.group(1)) if capacity_match else 0
                    total_tb = round(total_mb / (1024**2), 2)
                    
                    if status.upper() not in ["OK", "NORMAL"]:
                        has_error, status_text = True, "LỖI"
                    else:
                        status_text = "B.thường"
                    results.append(f"Disk-{i+1}: {status_text} ({total_tb}TB)")
                return " | ".join(results), has_error
        except Exception as e:
            return f"Lỗi kết nối API HDD: {str(e)}", True
        return "N/A", False

    @staticmethod
    def get_system_info(base_url, username, password):
        url = f"{base_url}/ISAPI/System/deviceInfo"
        try:
            response = requests.get(url, auth=HTTPDigestAuth(username, password), timeout=10)
            if response.status_code == 200:
                ver = re.search(r'<firmwareVersion>(.*?)</firmwareVersion>', response.text)
                return f"Online (Firmware: {ver.group(1).strip() if ver else 'N/A'})"
        except: pass
        return "Offline/Lỗi kết nối"
    

def send_alert_email():
    # Nếu không có gì để báo cáo thì thôi
    if not ALL_HDD_STATUS and not NG_CAMERAS:
        print("Không có dữ liệu hệ thống để gửi mail.")
        return

    msg = MIMEMultipart()
    msg["From"] = GMAIL_USER
    msg["To"] = ", ".join(GMAIL_TO)
    
    # Tùy biến tiêu đề mail: Có lỗi thì Báo đỏ, Không lỗi thì Báo xanh
    if NG_CAMERAS or SYSTEM_ALERTS:
        msg["Subject"] = f"🔴 [CẢNH BÁO] Phát hiện {len(SYSTEM_ALERTS)} NVR lỗi & {len(NG_CAMERAS)} Camera lỗi"
    else:
        msg["Subject"] = "🟢 [BÁO CÁO] Hệ thống Camera hoạt động bình thường 100%"

    # ==========================================
    # XÂY DỰNG GIAO DIỆN BẢNG HTML CHO EMAIL
    # ==========================================
    html_body = f"""
    <html>
    <head>
        <style>
            body {{ font-family: Arial, sans-serif; color: #333; }}
            table {{ border-collapse: collapse; width: 100%; margin-bottom: 20px; }}
            th, td {{ border: 1px solid #dddddd; text-align: left; padding: 10px; font-size: 14px; }}
            th {{ background-color: #f4f4f4; color: #000; font-weight: bold; }}
            .status-ok {{ color: #008000; font-weight: bold; }}
            .status-ng {{ color: #d93025; font-weight: bold; }}
            .header-box {{ background-color: #004a99; color: white; padding: 15px; text-align: center; margin-bottom: 20px; }}
            h3 {{ color: #004a99; border-bottom: 2px solid #004a99; padding-bottom: 5px; display: inline-block; }}
        </style>
    </head>
    <body>
        <div class="header-box">
            <h2 style="margin: 0;">BÁO CÁO TỔNG HỢP HỆ THỐNG CAMERA</h2>
            <p style="margin: 5px 0 0 0;">Thời gian quét: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}</p>
        </div>

        <h3>1. TÌNH TRẠNG ĐẦU GHI & Ổ CỨNG (NVR / HDD)</h3>
        <table>
            <tr>
                <th width="5%">STT</th>
                <th width="20%">Khu vực / Xưởng</th>
                <th width="10%">Hãng</th>
                <th width="25%">Trạng thái Hệ thống</th>
                <th width="40%">Sức khỏe Ổ cứng (HDD)</th>
            </tr>
    """

    # Bóc tách chuỗi ALL_HDD_STATUS để điền vào bảng
    if ALL_HDD_STATUS:
        for i, hdd_str in enumerate(ALL_HDD_STATUS):
            # Dùng regex để tách chuỗi kiểu "[Xưởng] (Hãng) - Info | HDD: detail"
            match = re.match(r'\[(.*?)\]\s+\((.*?)\)\s+-\s+(.*?)\s+\|\s+HDD:\s+(.*)', hdd_str)
            if match:
                factory, brand, sys_info, hdd_info = match.groups()
            else:
                factory, brand, sys_info, hdd_info = "Unknown", "Unknown", hdd_str, "Unknown"

            # Phân loại màu sắc (Xanh / Đỏ)
            hdd_class = "status-ng" if "LỖI" in hdd_info.upper() else "status-ok"
            sys_class = "status-ng" if "Offline" in sys_info else "status-ok"

            html_body += f"""   
                <tr>
                    <td align="center">{i+1}</td>
                    <td><b>{factory}</b></td>
                    <td align="center">{brand}</td>
                    <td class="{sys_class}">{sys_info}</td>
                    <td class="{hdd_class}">{hdd_info}</td>
                </tr>
            """
    else:
        html_body += """<tr><td colspan="5" align="center">Không có dữ liệu kiểm tra thiết bị.</td></tr>"""

    html_body += """
        </table>

        <h3>2. DANH SÁCH CAMERA MẤT HÌNH ẢNH (NG)</h3>
    """

    # Bảng danh sách Camera lỗi
    if NG_CAMERAS:
        html_body += """
        <table>
            <tr>
                <th width="5%">STT</th>
                <th width="25%">Xưởng / Vị trí</th>
                <th width="35%">Tên Camera (Kênh)</th>
                <th width="35%">Mô tả lỗi thuật toán AI</th>
            </tr>
        """
        for i, item in enumerate(NG_CAMERAS):
            html_body += f"""
                <tr>
                    <td align="center">{i+1}</td>
                    <td><b>{item.get('Xưởng', 'Unknown')}</b></td>
                    <td class="status-ng">{item.get('Kênh', 'Unknown')}</td>
                    <td>{item.get('reason', '')}</td>
                </tr>
            """
        html_body += """
        </table>
        <p style="color: #666; font-style: italic;">* Lưu ý: Đã đính kèm ảnh chụp các camera bị lỗi bên dưới email này để đối chiếu.</p>
        """
    else:
        html_body += """<p class="status-ok" style="font-size: 16px;">✅ Tuyệt vời! Tất cả camera đều hiển thị hình ảnh bình thường, không phát hiện nhiễu hay sọc đen.</p>"""

    html_body += """
    </body>
    </html>
    """

    # Đính kèm nội dung HTML vào Mail
    msg.attach(MIMEText(html_body, "html"))

    # ==========================================
    # PHẦN 3: ĐÍNH KÈM ẢNH CAMERA LỖI (Giữ nguyên)
    # ==========================================
    for item in NG_CAMERAS:
        if item.get('path') and os.path.exists(item['path']):
            try:
                with open(item["path"], "rb") as f:
                    part = MIMEBase("application", "octet-stream")
                    part.set_payload(f.read())
                encoders.encode_base64(part)
                part.add_header("Content-Disposition", f"attachment; filename={os.path.basename(item['path'])}")
                msg.attach(part)
            except: pass

    # Gửi mail
    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(GMAIL_USER, GMAIL_APP_PASSWORD)
            server.sendmail(GMAIL_USER, GMAIL_TO, msg.as_string())
        print("\n>>> Đã gửi email báo cáo!")
    except Exception as e:
        print(f"\n>>> Lỗi khi gửi email: {e}")


def main():
    # Thư mục chứa dữ liệu
    base_data_dir = r"D:\Automation\src\tasks\CAMERA\data"
    snapshot_dir = r"D:\Automation\src\tasks\CAMERA\data\snapshots"
    os.makedirs(snapshot_dir, exist_ok=True)

    # Khởi tạo Excel
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Camera Analysis"
    sheet.append(["Khu vực/Xưởng", "Tên Kênh", "Trạng Thái", "Thông tin HDD (Sức khỏe)", "Hệ thống", "Mô tả lỗi", "Ảnh minh họa"])

    # DANH SÁCH TẠM GIAM: Chứa các camera chụp thất bại ở Vòng 1
    SUSPICIOUS_CAMERAS = [] 

    # =========================================================================
    # VÒNG 1: QUÉT NHANH TOÀN BỘ HỆ THỐNG
    # =========================================================================
    print("=== BẮT ĐẦU VÒNG 1: QUÉT HỆ THỐNG DAHUA ===")
    camera_info_path = os.path.join(base_data_dir, "CameraInfo.txt")
    if os.path.exists(camera_info_path):
        with open(camera_info_path, "r") as file:
            for line in file.readlines():
                if not line.strip() or '|' not in line: continue
                factory_name, base_url, username, password = line.strip().split('|')
                print(f"\nChecking Dahua: {factory_name} ({base_url})")

                hdd_text, is_hdd_error = CameraUtility.get_hdd_status(base_url, username, password)
                sys_info = CameraUtility.get_system_info(base_url, username, password)
                ALL_HDD_STATUS.append(f"[{factory_name}] (DAHUA) - {sys_info} | HDD: {hdd_text}")
                if is_hdd_error: SYSTEM_ALERTS.append(f"{factory_name} (DAHUA): {hdd_text}")
                sheet.append([factory_name, "DEVICE_ROOT", "INFO", hdd_text, sys_info, "Thông tin đầu ghi", ""])

                channel_info = CameraUtility.get_channel_title_config(base_url, username, password)
                if not channel_info: continue

                for channel_id, channel_name in channel_info.items():
                    if "not in use" in channel_name.lower(): continue
                    snapshot_path = os.path.join(snapshot_dir, f"{factory_name}_{channel_id}.jpg")
                    
                    # CỐ GẮNG CHỤP VÀ ĐÁNH GIÁ (VÒNG 1)
                    if CameraUtility.get_snapshot(base_url, username, password, channel_id, snapshot_path):
                        # TRƯỜNG HỢP 1: KẾT NỐI ĐƯỢC -> KIỂM TRA AI
                        result = classify_image(snapshot_path)
                        
                        if not result["error"]:
                            # Tốt -> Ghi Excel OK
                            sheet.append([factory_name, channel_name, "OK", "-", "-", "", ""])
                            # (Đoạn chèn ảnh vào Excel giữ nguyên...)
                        else:
                            # LỖI AI (Đen/Sọc/Nhiễu) -> CHỐT LỖI NG NGAY LẬP TỨC
                            print(f"[-] LỖI AI: {factory_name} - {channel_name} ({result['description']})")
                            sheet.append([factory_name, channel_name, "NG", "-", "-", result["description"], ""])
                            # Đưa thẳng vào danh sách gửi mail cho sếp
                            NG_CAMERAS.append({
                                "path": snapshot_path, 
                                "Xưởng": factory_name, 
                                "Kênh": channel_name, 
                                "reason": result["description"]
                            })
                    else:
                        # TRƯỜNG HỢP 2: LỖI KẾT NỐI -> ĐƯA VÀO DIỆN NGHI NGỜ ĐỂ ĐỢI 5 PHÚT
                        print(f"[?] MẤT KẾT NỐI: {factory_name} - {channel_name} -> Chờ kiểm tra lại sau 5p")
                        SUSPICIOUS_CAMERAS.append({
                            "brand": "Dahua", # Hoặc "Hikvision" tùy vòng lặp
                            "factory": factory_name, 
                            "url": base_url, 
                            "user": username, 
                            "pass": password, 
                            "id": channel_id, 
                            "name": channel_name, 
                            "path": snapshot_path
                        })

    print("\n=== BẮT ĐẦU VÒNG 1: QUÉT HỆ THỐNG HIKVISION ===")
    hik_info_path = os.path.join(base_data_dir, "HikvisionInfo.txt")
    if os.path.exists(hik_info_path):
        with open(hik_info_path, "r") as file:
            for line in file.readlines():
                if not line.strip() or '|' not in line: continue
                factory_name, base_url, username, password = line.strip().split('|')
                print(f"\nChecking Hikvision: {factory_name} ({base_url})")

                hdd_text, is_hdd_error = HikvisionUtility.get_hdd_status(base_url, username, password)
                sys_info = HikvisionUtility.get_system_info(base_url, username, password)
                ALL_HDD_STATUS.append(f"[{factory_name}] (HIKVISION) - {sys_info} | HDD: {hdd_text}")
                if is_hdd_error: SYSTEM_ALERTS.append(f"{factory_name} (HIK): {hdd_text}")
                sheet.append([factory_name, "DEVICE_ROOT", "INFO", hdd_text, sys_info, "Thông tin đầu ghi", ""])

                channel_info = HikvisionUtility.get_channel_title_config(base_url, username, password)
                if not channel_info: continue

                for channel_id, channel_name in channel_info.items():
                    if "not in use" in channel_name.lower(): continue
                    snapshot_path = os.path.join(snapshot_dir, f"HIK_{factory_name}_{channel_id}.jpg")
                    
                    # CỐ GẮNG CHỤP VÀ ĐÁNH GIÁ (VÒNG 1)
                    if HikvisionUtility.get_snapshot(base_url, username, password, channel_id, snapshot_path):
                        result = classify_image(snapshot_path)
                        if not result["error"]:
                            # CAMERA TỐT -> Ghi Excel luôn
                            print(f"{factory_name} - {channel_name}: OK")
                            sheet.append([factory_name, channel_name, "OK", "-", "-", "", ""])
                            current_row = sheet.max_row
                            try:
                                img = Image(snapshot_path)
                                img.width, img.height = 240, 135
                                sheet.add_image(img, f"G{current_row}")
                                sheet.column_dimensions["G"].width = img.width / 7
                                sheet.row_dimensions[current_row].height = img.height
                            except: pass
                        else:
                            # LỖI AI (Đen/Sọc/Nhiễu) -> CHỐT LỖI NG NGAY LẬP TỨC
                            print(f"[-] LỖI AI: {factory_name} - {channel_name} ({result['description']})")
                            sheet.append([factory_name, channel_name, "NG", "-", "-", result["description"], ""])
                            # Đưa thẳng vào danh sách gửi mail cho sếp
                            NG_CAMERAS.append({
                                "path": snapshot_path, 
                                "Xưởng": factory_name, 
                                "Kênh": channel_name, 
                                "reason": result["description"]
                            })
                    else:
                        # MẤT MẠNG -> Đưa vào diện nghi ngờ
                        print(f"[?] NGHI NGỜ: {factory_name} - {channel_name} (Mất kết nối)")
                        SUSPICIOUS_CAMERAS.append({"brand": "Hikvision", "factory": factory_name, "url": base_url, "user": username, "pass": password, "id": channel_id, "name": channel_name, "path": snapshot_path})

    # =========================================================================
    # THỜI GIAN TRỄ LŨY TIẾN (Đợi mạng phục hồi)
    # =========================================================================
    if SUSPICIOUS_CAMERAS:
        print("\n==========================================================")
        print(f"[*] CẢNH BÁO: Phát hiện {len(SUSPICIOUS_CAMERAS)} camera có dấu hiệu NG.")
        print("[*] Đang kích hoạt cơ chế chờ 5 phút (300 giây) để mạng phục hồi...")
        print("==========================================================")
        time.sleep(300) # Nghỉ 5 phút chẵn

        # =========================================================================
        # VÒNG 2: CHỐT HẠ (Xác nhận chết hay đã sống lại)
        # =========================================================================
        print("\n=== BẮT ĐẦU VÒNG 2: KIỂM TRA LẠI CAMERA NGHI NGỜ ===")
        for cam in SUSPICIOUS_CAMERAS:
            print(f"Re-checking {cam['factory']} - {cam['name']}...")
            
            # Khởi tạo mặc định
            is_success = False
            final_reason = ""
            
            # Gọi lệnh chụp ảnh lần nữa
            if cam["brand"] == "Dahua":
                success_snap = CameraUtility.get_snapshot(cam["url"], cam["user"], cam["pass"], cam["id"], cam["path"])
            else:
                success_snap = HikvisionUtility.get_snapshot(cam["url"], cam["user"], cam["pass"], cam["id"], cam["path"])

            # Đánh giá kết quả Vòng 2
            if success_snap:
                result = classify_image(cam["path"])
                if not result["error"]:
                    is_success = True
                else:
                    final_reason = result["description"]
            else:
                final_reason = "Mất kết nối/Timeout (Đã rớt mạng > 5 phút)"

            # RA QUYẾT ĐỊNH CUỐI CÙNG
            if is_success:
                print(f"-> ✅ ĐÃ PHỤC HỒI MẠNG: {cam['factory']} - {cam['name']}")
                sheet.append([cam["factory"], cam["name"], "OK", "-", "-", "Đã phục hồi sau khi rớt mạng", ""])
                current_row = sheet.max_row
                try:
                    img = Image(cam["path"])
                    img.width, img.height = 240, 135
                    sheet.add_image(img, f"G{current_row}")
                except: pass
            else:
                print(f"-> ❌ CHẾT THẬT SỰ: {cam['factory']} - {cam['name']} ({final_reason})")
                sheet.append([cam["factory"], cam["name"], "NG", "-", "-", final_reason, ""])
                current_row = sheet.max_row
                try:
                    img = Image(cam["path"])
                    img.width, img.height = 240, 135
                    sheet.add_image(img, f"G{current_row}")
                except: pass
                
                # CHỈ ĐƯA VÀO DANH SÁCH GỬI MAIL KHI ĐÃ THẤT BẠI Ở VÒNG 2
                NG_CAMERAS.append({"path": cam["path"] if os.path.exists(cam["path"]) else None, "Xưởng": cam["factory"], "Kênh": cam["name"], "reason": final_reason})

    # =========================================================================
    # LƯU EXCEL VÀ GỬI MAIL
    # =========================================================================
    current_date = datetime.now().strftime("%Y%m%d")
    output_path = fr"N:\Ｃ：ＮＩＶＳ総務・人事(GA&HR)\０４：ＩＴ\１７：RPA Report\1. Camera\CameraAnalysis_{current_date}.xlsx"
    workbook.save(output_path)
    
    send_alert_email()
    print(f"Results saved to {output_path}")

if __name__ == "__main__":
    main()